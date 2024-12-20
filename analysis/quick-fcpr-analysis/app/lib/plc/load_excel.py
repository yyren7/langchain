import openpyxl


def getParamDict(book_name: str, sheet_name: str) -> dict:
    REF_COL_NO = 2
    MAX_COL_NO = 2
    wb = openpyxl.load_workbook(book_name, data_only=True)  # データオンリー
    sheet = wb[sheet_name]
    dict_data = {}
    # 行走査
    for row_i in range(1, sheet.max_row + 1):
        cell_1 = sheet.cell(row=row_i, column=REF_COL_NO)
        # 値があれば
        if cell_1.value is not None:
            # 列走査
            for col_i in range(1, MAX_COL_NO):
                cell_2 = sheet.cell(row=row_i, column=REF_COL_NO + col_i)
                # 値があれば
                if cell_2.value is not None:
                    # bit用連想配列作成
                    dict_data[cell_1.value] = cell_2.value
    # print(dict_data)
    wb.close()
    return dict_data


def getWdataDict(book_name: str, sheet_name: str) -> dict:
    REF_ROW_NO = 3
    REF_COL_NO = 3
    MAX_COL_NO = 16
    wb = openpyxl.load_workbook(book_name, data_only=True)
    sheet = wb[sheet_name]
    dict_data = {}
    # 行走査
    for row_i in range(0, sheet.max_row):
        cell = sheet.cell(row=row_i + REF_ROW_NO, column=REF_COL_NO)
        # 値があれば
        if cell.value is not None:
            # word用連想配列作成
            dict_data[cell.value] = row_i
            # 列走査
            for col_i in range(0, MAX_COL_NO):
                cell = sheet.cell(row=row_i + REF_ROW_NO, column=REF_COL_NO + col_i)
                # 値があれば
                if (cell.value is not None) and (col_i != 0):
                    # bit用連想配列作成
                    dict_data[cell.value] = col_i - 1
    # print(dict_data)
    wb.close()
    return dict_data


def getDdataDict(book_name: str, sheet_name: str) -> dict:
    REF_ROW_NO = 3
    REF_COL_NO = 3
    wb = openpyxl.load_workbook(book_name, data_only=True)
    sheet = wb[sheet_name]
    dict_data = {}
    # 行走査
    for row_i in range(0, sheet.max_row):
        cell = sheet.cell(row=row_i + REF_ROW_NO, column=REF_COL_NO)
        # 値があれば
        if cell.value is not None:
            # word用連想配列作成
            dict_data[cell.value] = row_i
    # print(dict_data)
    wb.close()
    return dict_data


# NOTE: 追加
def getIdTable(book_name: str, sheet_name: str = 'IDTABLE') -> list:
    REF_ROW_NO = 3
    REF_COL_NO = 2
    # NOTE:データオンリー
    wb = openpyxl.load_workbook(book_name, data_only=True)
    sheet = wb[sheet_name]
    data_list = []
    # NOTE:行走査
    for row_i in range(REF_ROW_NO, sheet.max_row + 1):
        program_no = sheet.cell(row=row_i, column=REF_COL_NO).value
        config_no = sheet.cell(row=row_i, column=REF_COL_NO + 1).value
        model_no = sheet.cell(row=row_i, column=REF_COL_NO + 2).value
        model_type = sheet.cell(row=row_i, column=REF_COL_NO + 3).value
        model_name = sheet.cell(row=row_i, column=REF_COL_NO + 4).value
        if program_no is None:
            # NOTE: Noneであればbreak
            break
        data_list.append(
            dict(config=config_no,
                 program=program_no,
                 model=model_no,
                 mdl_type=model_type,
                 mdl_name=model_name)
        )

    wb.close()
    return data_list

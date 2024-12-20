import openpyxl

def getParamDict(book_name: str, sheet_name: str) -> dict:
    REF_COL_NO = 2
    MAX_COL_NO = 2
    wb = openpyxl.load_workbook(book_name, data_only=True) #データオンリー    
    sheet = wb[sheet_name]   
    dict_data = {}
    # 行走査
    for row_i in range(1,sheet.max_row+1):
        cell_1 = sheet.cell(row=row_i, column=REF_COL_NO)
        # 値があれば
        if(cell_1.value != None):
            # 列走査
            for col_i in range(1, MAX_COL_NO):
                cell_2 = sheet.cell(row=row_i, column=REF_COL_NO+col_i)
                # 値があれば
                if(cell_2.value != None):
                    # bit用連想配列作成
                    dict_data[cell_1.value] = cell_2.value
        
    # print(dict_data)                
    wb.close()
    return dict_data

def getWdataDict(book_name: str, sheet_name: str) -> dict:
    REF_ROW_NO = 3
    REF_COL_NO = 3
    MAX_COL_NO = 16
    wb = openpyxl.load_workbook(book_name, data_only=True)
    sheet = wb[sheet_name]    
    dict_data = {}
    # 行走査
    for row_i in range(0,sheet.max_row):
        cell = sheet.cell(row=row_i+REF_ROW_NO, column=REF_COL_NO)
        # 値があれば
        if(cell.value != None):
            # word用連想配列作成
            dict_data[cell.value] = row_i
            # 列走査
            for col_i in range(0, MAX_COL_NO):
                cell = sheet.cell(row=row_i+REF_ROW_NO, column=REF_COL_NO+col_i)
                # 値があれば
                if((cell.value != None) and not(col_i == 0)):
                    # bit用連想配列作成
                    dict_data[cell.value] = col_i - 1
    # print(dict_data)                
    wb.close()
    return dict_data

def getDdataDict(book_name: str, sheet_name: str) -> dict:
    REF_ROW_NO = 3
    REF_COL_NO = 3
    wb = openpyxl.load_workbook(book_name, data_only=True)
    sheet = wb[sheet_name]    
    dict_data = {}
    # 行走査
    for row_i in range(0,sheet.max_row):
        cell = sheet.cell(row=row_i+REF_ROW_NO, column=REF_COL_NO)
        # 値があれば
        if(cell.value != None):
            # word用連想配列作成
            dict_data[cell.value] = row_i
    # print(dict_data)
    wb.close()
    return dict_data

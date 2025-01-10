# ===============================================================================
# Name      : read_excel.py
# Version   : 1.0.0
# Brief     :
# Time-stamp: 2023-05-06 16:09
# Copyirght 2021 Hiroya Aoyama [aoyama.hiroya@nidec.com]
# ===============================================================================
import openpyxl
import io


def getHeaderAddress(book_name: str, sheet_name: str) -> dict:
    """指定されたExcelファイルとシートから先頭アドレスを取得する"""
    REF_ROW_NO = 8
    REF_COL_NO = 2

    with open(book_name, "rb") as f:
        in_mem_file = io.BytesIO(f.read())

    wb = openpyxl.load_workbook(in_mem_file, read_only=True,
                                keep_vba=True, data_only=True)
    sheet = wb[sheet_name]
    in_bit = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO).value
    in_word = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 1).value
    out_bit = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 2).value
    out_word = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 3).value
    wb.close()
    address_data = dict(IN_BIT_NO=in_bit,
                        IN_WORD_NO=in_word,
                        OUT_BIT_NO=out_bit,
                        OUT_WORD_NO=out_word,
                        )
    return address_data


def getBitData(book_name: str, sheet_name: str) -> dict:
    """Bitのキーワードを取得"""
    REF_ROW_NO = 3
    REF_COL_NO = 3
    MAX_COL_NO = 16

    with open(book_name, "rb") as f:
        in_mem_file = io.BytesIO(f.read())

    wb = openpyxl.load_workbook(in_mem_file, read_only=True,
                                keep_vba=True, data_only=True)
    sheet = wb[sheet_name]
    dict_data = {}
    # 行走査
    for row_i in range(0, sheet.max_row):
        cell = sheet.cell(row=row_i + REF_ROW_NO, column=REF_COL_NO)
        # 値があれば
        if cell.value is not None:
            # word用連想配列作成(Word要タグのID)
            dict_data[cell.value] = row_i
            # 列走査
            for col_i in range(0, MAX_COL_NO):
                cell = sheet.cell(row=row_i + REF_ROW_NO, column=REF_COL_NO + col_i)
                # 値があれば
                if (cell.value is not None) and (col_i != 0):
                    # bit用連想配列作成
                    dict_data[cell.value] = col_i - 1
    wb.close()
    return dict_data


def getWordData(book_name: str, sheet_name: str) -> dict:
    """Wordのキーワードを取得"""
    REF_ROW_NO = 3
    REF_COL_NO = 3

    with open(book_name, "rb") as f:
        in_mem_file = io.BytesIO(f.read())

    wb = openpyxl.load_workbook(in_mem_file, read_only=True,
                                keep_vba=True, data_only=True)
    sheet = wb[sheet_name]
    dict_data = {}
    # 行走査
    for row_i in range(0, sheet.max_row):
        cell = sheet.cell(row=row_i + REF_ROW_NO, column=REF_COL_NO)
        # 値があれば
        if cell.value is not None:
            # word用連想配列作成
            dict_data[cell.value] = row_i
    wb.close()
    return dict_data


def getPLCParam(book_name: str, sheet_name: str) -> dict:
    """PLCの通信設定を取得"""
    REF_ROW_NO = 4
    REF_COL_NO = 2

    with open(book_name, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
    wb = openpyxl.load_workbook(in_mem_file, read_only=True,
                                keep_vba=True, data_only=True)
    sheet = wb[sheet_name]
    # NOTE:行走査
    ip_address = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO).value
    port = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 1).value
    manufacturer = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 2).value
    series = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 3).value
    plc_protocol = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 4).value
    transport_protocol = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 5).value
    bit = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 6).value
    word = sheet.cell(row=REF_ROW_NO, column=REF_COL_NO + 7).value

    plc_param = dict(ip=ip_address,
                     port=str(port),
                     manufacturer=manufacturer,
                     series=series,
                     plc_protocol=plc_protocol,
                     transport_protocol=transport_protocol,
                     bit=bit,
                     word=word,
                     double_word='')
    wb.close()
    return plc_param


def getParamDict(book_name: str, sheet_name: str) -> dict:
    """RangeSheet読込"""
    # NOTE: NMOJ 用
    REF_COL_NO = 2
    MAX_COL_NO = 2
    wb = openpyxl.load_workbook(book_name, read_only=True,
                                keep_vba=True, data_only=True)
    sheet = wb[sheet_name]
    dict_data = {}
    # 行走査
    for row_i in range(1, sheet.max_row + 1):
        cell_1 = sheet.cell(row=row_i, column=REF_COL_NO)
        # 値があれば
        if cell_1.value is not None:
            # 列走査
            for col_i in range(1, MAX_COL_NO):
                cell_2 = sheet.cell(row=row_i, column=REF_COL_NO + col_i)
                # 値があれば
                if cell_2.value is not None:
                    # bit用連想配列作成
                    dict_data[cell_1.value] = cell_2.value
    wb.close()
    return dict_data


def getIdTable(book_name: str, sheet_name: str = 'IDTABLE') -> list:
    # NOTE: NMOJ用
    REF_ROW_NO = 3
    REF_COL_NO = 2
    # NOTE:データオンリー
    wb = openpyxl.load_workbook(book_name, read_only=True,
                                keep_vba=True, data_only=True)
    sheet = wb[sheet_name]
    data_list = []
    # NOTE:行走査
    for row_i in range(REF_ROW_NO, sheet.max_row + 1):
        program_no = sheet.cell(row=row_i, column=REF_COL_NO).value
        config_no = sheet.cell(row=row_i, column=REF_COL_NO + 1).value
        model_no = sheet.cell(row=row_i, column=REF_COL_NO + 2).value
        model_type = sheet.cell(row=row_i, column=REF_COL_NO + 3).value
        model_name = sheet.cell(row=row_i, column=REF_COL_NO + 4).value
        if program_no is None:
            # NOTE: Noneであればbreak
            break
        data_list.append(
            dict(config=config_no,
                 program=program_no,
                 model=model_no,
                 mdl_type=model_type,
                 mdl_name=model_name)
        )

    wb.close()
    return data_list
